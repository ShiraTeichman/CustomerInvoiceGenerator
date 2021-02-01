
import os
import openpyxl as xl
from InvoiceGenerator.pdf import SimpleInvoice
from tempfile import NamedTemporaryFile
from InvoiceGenerator.api import Invoice, Item, Client, Provider, Creator

wb = xl.load_workbook('ShiraTeichman-People_Info.xlsx')
sheet = wb['Sheet1']


# choose english as language
os.environ["INVOICE_LANG"] = "en"
for row in range(2, sheet.max_row + 1):
    paid = sheet.cell(row, 3)
    if not paid:
        client = Client(summary=f'Name: {sheet.cell(row, 1)}', address=sheet.cell(row, 5), zip_code=sheet.cell(row, 8), city=sheet.cell(row, 6), vat_id=f'tax id: {sheet.cell(row, 4)}', email=sheet.cell(row, 2))
        provider = Provider('Shira Teichman', bank_account='2600420569', bank_code='2010')
        creator = Creator('John Doe')
        invoice = Invoice(client, provider, creator)
        invoice.currency_locale = 'en_US.UTF-8'
        invoice.add_item(Item(32, 600, description="Item 1"))
        invoice.add_item(Item(60, 50, description="Item 2", tax=21))
        invoice.add_item(Item(50, 60, description="Item 3", tax=0))
        invoice.add_item(Item(5, 600, description="Item 4", tax=15))
        pdf = SimpleInvoice(invoice)
        pdf.gen("invoice.pdf", generate_qr_code=True)
#making a pdf



